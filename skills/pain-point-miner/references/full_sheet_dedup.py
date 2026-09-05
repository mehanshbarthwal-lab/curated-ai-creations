import openpyxl
from collections import defaultdict
import sys
import re

# TARGET_PATH and SHEET_NAME should be updated per project.
FILE_PATH = r'F:\SVS Internship Project\consumer_problem_bank_updated.xlsx'

def extract_source_id(url):
    if not url:
        return None
    url = str(url).strip()
    m = re.search(r'/comments/([a-z0-9]+)', url)  # reddit thread id
    if m:
        return ('reddit', m.group(1))
    m = re.search(r'[?&]v=([a-zA-Z0-9_-]{6,})', url)  # youtube video id
    if m:
        return ('youtube', m.group(1))
    m = re.search(r'youtu\.be/([a-zA-Z0-9_-]{6,})', url)
    if m:
        return ('youtube', m.group(1))
    return ('other', url)

def run_dedup():
    wb = openpyxl.load_workbook(FILE_PATH)
    sheet = wb.active
    
    id_map = defaultdict(list)
    
    for r in range(2, sheet.max_row + 1):
        ph = sheet.cell(row=r, column=1).value
        if not ph: continue
        
        link = sheet.cell(row=r, column=9).value
        key = extract_source_id(link)
        
        if key:
            id_map[key].append({
                'row': r,
                'ph': ph,
                'link': link,
                'quote': sheet.cell(row=r, column=6).value
            })
            
    collisions = {k: v for k, v in id_map.items() if len(v) > 1}
    
    if collisions:
        print("COLLISIONS DETECTED:")
        for tid, items in collisions.items():
            print(f"\nSource ID: {tid}")
            for item in items:
                print(f"  Row {item['row']} ({item['ph']}) - {item['link']}")
                print(f"    Quote: {item['quote']}")
        print("\nReview the collisions above. For any group of 2+ rows sharing the same source ID, read and compare the actual quotes/content directly (in the original language if translated) to confirm whether they're genuinely distinct comments or the same underlying text logged more than once.")
        sys.exit(1)
    else:
        print(f"Full-sheet dedup complete. 0 collisions found across {sheet.max_row - 1} entries.")
        sys.exit(0)

if __name__ == '__main__':
    run_dedup()
